import boto3 
import pprint
from openpyxl import load_workbook
AWS_REGION = "ap-southeast-2"
ec2 = boto3.client('ec2', region_name = AWS_REGION)

wb = load_workbook("myworkbook".xlsx")  
ws = wb['Sheet1']
column = ws['D'][0:3089]  # work book has 3089 rows to review
column_list = [column[x].value for x in range(len(column))]  

#snapshot_string =" ".join(map(str,column_list)).split(",") #USE MAP TO CONVERT LIST TO STRING # USE SPLIT METHOD TO DIVIDE STRINGS AS MULTIPLE SUBSSTRINGS # 4/11 Unneeded
pprint.pprint(column_list)

print(f'The snapshot for loop is starting')
for snap in column_list:
    print(f"Snapshot {snap} are deleting")
    if ec2.delete_snapshot(SnapshotId=snap) and SnapshotId.InUse == "true":
        print(f'Snapshots are in use by different AMIs')
    else: 
        print(f"Snapshot {snap} have been deleted")


'''NOTE: 
SCRIPT WORKED
OUTPUT RETURNED An error occurred (InvalidSnapshot.InUse) when calling the DeleteSnapshot operation: The snapshot snap-09d9a33133b8dcff8 is currently in use by ami-092b734a6f7e56317.

IMPROVEMENT OPPORUNITY: IF STATEMENT NOT WORKING
'''
