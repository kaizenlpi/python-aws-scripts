# Purpose: Cost Optimization

import boto3 
import pprint
from openpyxl import load_workbook
AWS_REGION = "us-gov-west-1"
ec2 = boto3.client('ec2', region_name = AWS_REGION)

wb = load_workbook("CREPCE.xlsx")  
ws = wb['Sheet1']  
column = ws['D']  
column_list = [column[x].value for x in range(len(column))]  
for volume in volumes_string:
    print(f"Volumes {volume} are deleting")
    try:
        ec2.delete_volume(VolumeId=volume)
    except Exception as e:
        print(e)
        continue
    print(f"Volumes {volume} have been deleted")
